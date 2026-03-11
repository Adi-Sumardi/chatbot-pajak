"""OCR service for extracting bank statement tables from PDF files using pdfplumber."""

import asyncio
import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.config import get_settings

logger = logging.getLogger(__name__)
settings = get_settings()

BANK_DETECTION_PATTERNS = {
    "bca": ["bank central asia", "bca", "pt bank central asia", "kcu bca", "cabang bca"],
    "mandiri": ["bank mandiri", "pt bank mandiri", "mandiri", "livin' by mandiri"],
    "bri": ["bank rakyat indonesia", "bri", "pt bank rakyat indonesia", "britama", "simpedes"],
    "bni": ["bank negara indonesia", "bni", "pt bank negara indonesia", "bni taplus"],
    "cimb": ["cimb niaga", "pt bank cimb niaga", "cimb"],
    "bsi": ["bank syariah indonesia", "bsi", "pt bank syariah indonesia"],
    "danamon": ["bank danamon", "danamon", "pt bank danamon"],
    "permata": ["bank permata", "permata", "pt bank permata"],
    "ocbc": ["ocbc nisp", "ocbc", "pt bank ocbc nisp"],
    "mega": ["bank mega", "mega", "pt bank mega"],
}

# Bank-specific table extraction settings for pdfplumber
BANK_TABLE_SETTINGS = {
    "bca": {
        "vertical_strategy": "text",
        "horizontal_strategy": "text",
        "snap_x_tolerance": 5,
        "snap_y_tolerance": 5,
        "join_x_tolerance": 5,
        "join_y_tolerance": 5,
        "min_words_vertical": 2,
        "min_words_horizontal": 1,
        "text_x_tolerance": 3,
        "text_y_tolerance": 3,
    },
    "mandiri": {
        "vertical_strategy": "text",
        "horizontal_strategy": "text",
        "snap_x_tolerance": 5,
        "snap_y_tolerance": 5,
        "join_x_tolerance": 5,
        "join_y_tolerance": 5,
        "min_words_vertical": 2,
        "min_words_horizontal": 1,
        "text_x_tolerance": 3,
        "text_y_tolerance": 3,
    },
    "default": {
        "vertical_strategy": "text",
        "horizontal_strategy": "text",
        "snap_x_tolerance": 5,
        "snap_y_tolerance": 5,
        "join_x_tolerance": 8,
        "join_y_tolerance": 5,
        "min_words_vertical": 2,
        "min_words_horizontal": 1,
        "text_x_tolerance": 3,
        "text_y_tolerance": 3,
    },
}

# Indonesian month names for date parsing
BULAN_MAP = {
    "jan": 1, "januari": 1, "feb": 2, "februari": 2, "mar": 3, "maret": 3,
    "apr": 4, "april": 4, "mei": 5, "may": 5, "jun": 6, "juni": 6,
    "jul": 7, "juli": 7, "agu": 8, "agustus": 8, "aug": 8,
    "sep": 9, "september": 9, "okt": 10, "oktober": 10, "oct": 10,
    "nov": 11, "november": 11, "des": 12, "desember": 12, "dec": 12,
}


def detect_bank_from_pdf(file_path: str) -> str | None:
    """Auto-detect bank name from PDF text content."""
    try:
        with pdfplumber.open(file_path) as pdf:
            text = ""
            for page in pdf.pages[:2]:
                page_text = page.extract_text()
                if page_text:
                    text += page_text.lower() + "\n"

            if not text:
                return None

            for bank_key, patterns in BANK_DETECTION_PATTERNS.items():
                for pattern in patterns:
                    if pattern in text:
                        return bank_key

        return None
    except Exception as e:
        logger.error(f"Bank detection error: {e}")
        return None


def _parse_amount(text: str | None) -> Decimal | None:
    """Parse Indonesian-format amount string to Decimal.

    Handles formats like:
    - 1.234.567,89  (dots=thousands, comma=decimal)
    - 1,234,567.89  (commas=thousands, dot=decimal)
    - 1234567.89
    - 1234567
    - (1.234,56) or -1.234,56  (negative/debit)
    - Rp 1.234.567
    - DB 1.234.567 / CR 1.234.567
    """
    if not text:
        return None

    cleaned = text.strip()

    # Skip empty / placeholder values
    if cleaned in ("", "-", "--", "0", "0,00", "0.00", ".", ","):
        return None

    # Remove currency prefix/suffix
    cleaned = re.sub(r"(?i)^(rp\.?\s*|idr\s*)", "", cleaned)
    cleaned = re.sub(r"(?i)\s*(db|cr|dr)\.?\s*$", "", cleaned)
    cleaned = re.sub(r"(?i)^(db|cr|dr)\.?\s*", "", cleaned)

    # Check for negative indicators: parentheses or leading minus
    is_negative = False
    if cleaned.startswith("(") and cleaned.endswith(")"):
        is_negative = True
        cleaned = cleaned[1:-1].strip()
    elif cleaned.startswith("-"):
        is_negative = True
        cleaned = cleaned[1:].strip()

    # Remove spaces
    cleaned = cleaned.replace(" ", "")

    # Determine format: Indonesian (1.234,56) vs English (1,234.56)
    last_dot = cleaned.rfind(".")
    last_comma = cleaned.rfind(",")

    if last_dot > -1 and last_comma > -1:
        if last_comma > last_dot:
            # Format: 1.234.567,89 (Indonesian) - dot=thousand, comma=decimal
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            # Format: 1,234,567.89 (English) - comma=thousand, dot=decimal
            cleaned = cleaned.replace(",", "")
    elif last_comma > -1:
        # Only commas: could be 1,234,567 (thousands) or 1234,56 (decimal)
        parts = cleaned.split(",")
        if len(parts) == 2 and len(parts[-1]) <= 2:
            # Likely decimal: 1234,56
            cleaned = cleaned.replace(",", ".")
        else:
            # Likely thousands: 1,234,567
            cleaned = cleaned.replace(",", "")
    elif last_dot > -1:
        # Only dots: could be 1.234.567 (thousands) or 1234.56 (decimal)
        parts = cleaned.split(".")
        if len(parts) > 2:
            # Multiple dots = thousand separators: 1.234.567
            cleaned = cleaned.replace(".", "")
        elif len(parts) == 2 and len(parts[-1]) == 3 and len(parts[0]) <= 3:
            # Ambiguous but likely thousands for amounts like 100.000
            cleaned = cleaned.replace(".", "")
        # else: single dot with 1-2 decimals, keep as is

    # Remove any remaining non-numeric chars except dot
    cleaned = re.sub(r"[^\d.]", "", cleaned)

    if not cleaned or cleaned == ".":
        return None

    try:
        val = Decimal(cleaned)
        return -val if is_negative else val
    except (InvalidOperation, ValueError):
        return None


def _parse_date(text: str | None, year: int | None = None) -> date | None:
    """Try parsing common Indonesian bank date formats."""
    if not text or not text.strip():
        return None
    text = text.strip()

    # Remove extra whitespace
    text = re.sub(r"\s+", " ", text)

    # Pattern: DD/MM/YYYY or DD-MM-YYYY or DD.MM.YYYY
    m = re.match(r"(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})", text)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass

    # Pattern: DD/MM/YY
    m = re.match(r"(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2})(?:\D|$)", text)
    if m:
        try:
            yr = int(m.group(3))
            yr = yr + 2000 if yr < 100 else yr
            return date(yr, int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass

    # Pattern: DD MMM YYYY or DD-MMM-YYYY (e.g. "15 Jan 2024", "15-Jan-2024")
    m = re.match(r"(\d{1,2})[\s\-.]([A-Za-z]{3,9})[\s\-.]?(\d{2,4})?", text)
    if m:
        try:
            day = int(m.group(1))
            month_str = m.group(2).lower()
            month = BULAN_MAP.get(month_str)
            if month:
                yr = int(m.group(3)) if m.group(3) else (year or datetime.now().year)
                if yr < 100:
                    yr += 2000
                return date(yr, month, day)
        except (ValueError, TypeError):
            pass

    # Pattern: DD/MM (no year)
    m = re.match(r"(\d{1,2})[/\-.](\d{1,2})(?:\D|$)", text)
    if m:
        try:
            day, month = int(m.group(1)), int(m.group(2))
            yr = year or datetime.now().year
            return date(yr, month, day)
        except ValueError:
            pass

    # Pattern: DDMM (4 digits, used by some banks like BCA)
    m = re.match(r"^(\d{4})$", text)
    if m:
        try:
            day, month = int(text[:2]), int(text[2:])
            if 1 <= day <= 31 and 1 <= month <= 12:
                yr = year or datetime.now().year
                return date(yr, month, day)
        except ValueError:
            pass

    return None


def extract_tables_from_pdf(
    file_path: str,
    bank_name: str | None = None,
) -> list[list[list[str | None]]]:
    """Extract all tables from a PDF using pdfplumber with optimized settings."""
    all_tables = []
    table_settings = BANK_TABLE_SETTINGS.get(bank_name or "", BANK_TABLE_SETTINGS["default"])

    with pdfplumber.open(file_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            # Try with explicit lines/edges first (works better for lined tables)
            tables = page.extract_tables(
                table_settings={
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                }
            )

            # If no tables found with lines strategy, try text-based
            if not tables:
                tables = page.extract_tables(table_settings=table_settings)

            # If still no tables, try explicit strategy
            if not tables:
                tables = page.extract_tables(
                    table_settings={
                        "vertical_strategy": "text",
                        "horizontal_strategy": "lines_strict",
                        "snap_x_tolerance": 10,
                        "text_x_tolerance": 5,
                    }
                )

            for table in tables:
                if table and len(table) > 1:
                    # Clean up cells: strip whitespace, normalize
                    cleaned_table = []
                    for row in table:
                        cleaned_row = []
                        for cell in row:
                            if cell is not None:
                                # Replace multiple spaces/newlines with single space
                                cell = re.sub(r"\s+", " ", str(cell)).strip()
                                if not cell:
                                    cell = None
                            cleaned_row.append(cell)
                        cleaned_table.append(cleaned_row)
                    all_tables.append(cleaned_table)

            logger.debug(f"Page {page_num + 1}: found {len(tables)} tables")

    return all_tables


def _detect_columns(header_row: list[str | None]) -> dict[str, int]:
    """Detect column indices from header row."""
    mapping = {}
    if not header_row:
        return mapping

    for i, cell in enumerate(header_row):
        if not cell:
            continue
        cell_lower = cell.strip().lower()

        # Date columns
        if any(k in cell_lower for k in ("tanggal", "tgl", "date", "valuta", "posting")):
            if "tanggal" not in mapping:
                mapping["tanggal"] = i
        # Description columns
        elif any(k in cell_lower for k in (
            "keterangan", "uraian", "deskripsi", "description", "remark",
            "transaksi", "transaction", "berita", "narrative"
        )):
            mapping["keterangan"] = i
        # Debit columns
        elif any(k in cell_lower for k in ("debet", "debit", "db", "withdrawal", "penarikan", "keluar")):
            mapping["debit"] = i
        # Credit columns
        elif any(k in cell_lower for k in ("kredit", "credit", "cr", "deposit", "setoran", "masuk")):
            mapping["kredit"] = i
        # Balance columns
        elif any(k in cell_lower for k in ("saldo", "balance")):
            mapping["saldo"] = i
        # Combined amount column
        elif any(k in cell_lower for k in ("mutasi", "amount", "jumlah", "nominal")):
            if "debit" not in mapping:
                mapping["debit"] = i
        # No/Nomor column - skip
        elif any(k in cell_lower for k in ("no", "nomor", "no.")):
            pass

    return mapping


def _is_header_or_noise(row: list[str | None], col_map: dict[str, int]) -> bool:
    """Check if a row is a repeated header, footer, or noise."""
    raw = " ".join(str(c or "").lower() for c in row)

    # Skip repeated headers
    header_keywords = [
        "tanggal", "keterangan", "debit", "kredit", "saldo", "balance",
        "date", "description", "withdrawal", "deposit",
    ]
    header_count = sum(1 for kw in header_keywords if kw in raw)
    if header_count >= 3:
        return True

    # Skip totals and summary rows
    skip_patterns = [
        r"\btotal\b", r"\bsubtotal\b", r"\bjumlah\b", r"\bsaldo awal\b",
        r"\bsaldo akhir\b", r"\bopening balance\b", r"\bclosing balance\b",
        r"\bpage\s+\d+", r"\bhalaman\s+\d+", r"\bcontinued\b", r"\bbersambung\b",
        r"\bpindahan\b",
    ]
    for pattern in skip_patterns:
        if re.search(pattern, raw):
            return True

    return False


def _is_amount_cell(text: str | None) -> bool:
    """Check if text looks like a monetary amount."""
    if not text:
        return False
    cleaned = re.sub(r"[^\d.,\-() ]", "", text.strip())
    return bool(re.search(r"\d", cleaned))


def parse_bank_statement(
    tables: list[list[list[str | None]]],
    bank_name: str | None = None,
    year: int | None = None,
) -> list[dict]:
    """Parse extracted tables into structured transaction rows."""
    results = []
    row_num = 0

    for table in tables:
        if not table or len(table) < 2:
            continue

        # Try to detect header from first few rows
        col_map = {}
        start_row = 0
        for check_row in range(min(3, len(table))):
            col_map = _detect_columns(table[check_row])
            if len(col_map) >= 3:  # Need at least 3 columns detected
                start_row = check_row + 1
                break

        # If no header detected, try guessing from column count
        if len(col_map) < 3:
            ncols = max((len(r) for r in table if r), default=0)
            if ncols >= 6:
                # Possibly: No | Tanggal | Keterangan | Debit | Kredit | Saldo
                col_map = {"tanggal": 1, "keterangan": 2, "debit": 3, "kredit": 4, "saldo": 5}
            elif ncols >= 5:
                col_map = {"tanggal": 0, "keterangan": 1, "debit": 2, "kredit": 3, "saldo": 4}
            elif ncols >= 4:
                col_map = {"tanggal": 0, "keterangan": 1, "debit": 2, "kredit": 3}
            elif ncols >= 3:
                col_map = {"tanggal": 0, "keterangan": 1, "debit": 2}
            start_row = 0

            # Try to find where actual data starts (skip header-like rows)
            for i in range(min(5, len(table))):
                row = table[i]
                if not row:
                    continue
                first_cell = str(row[0] or "").strip() if row else ""
                # If first cell looks like a date or number, data starts here
                if re.match(r"\d{1,2}[/\-.]", first_cell) or re.match(r"^\d{1,4}$", first_cell):
                    start_row = i
                    break
                # Check if this row has numeric values (likely data)
                numeric_count = sum(1 for c in row if c and re.search(r"\d{3,}", str(c)))
                if numeric_count >= 2:
                    start_row = i
                    break

        prev_keterangan = None
        for row in table[start_row:]:
            if not row or all(not cell or not str(cell).strip() for cell in row):
                continue

            # Skip headers/noise/totals
            if _is_header_or_noise(row, col_map):
                continue

            raw_text = " | ".join(str(c or "") for c in row)

            # Extract values using column mapping
            def _get(key: str) -> str | None:
                idx = col_map.get(key)
                if idx is not None and idx < len(row):
                    val = row[idx]
                    return str(val).strip() if val else None
                return None

            tanggal_str = _get("tanggal")
            keterangan = _get("keterangan")
            debit_str = _get("debit")
            kredit_str = _get("kredit")
            saldo_str = _get("saldo")

            tanggal = _parse_date(tanggal_str, year)
            debit = _parse_amount(debit_str)
            kredit = _parse_amount(kredit_str)
            saldo = _parse_amount(saldo_str)

            # Handle continuation rows (no date, just extra description text)
            if not tanggal and not debit and not kredit and not saldo and keterangan:
                # This is likely a continuation of the previous row's description
                if results and prev_keterangan is not None:
                    results[-1]["keterangan"] = (results[-1]["keterangan"] or "") + " " + keterangan.strip()
                    results[-1]["raw_text"] += " | " + raw_text
                continue

            # Skip rows with no useful data at all
            if not tanggal and not debit and not kredit and not saldo and not keterangan:
                continue

            # Skip sub-header labels
            if keterangan:
                ket_lower = keterangan.strip().lower()
                if ket_lower in (
                    "keterangan", "uraian", "description", "total",
                    "tanggal", "date", "saldo", "balance",
                ):
                    continue

            row_num += 1
            results.append({
                "row_number": row_num,
                "tanggal": tanggal,
                "keterangan": keterangan.strip() if keterangan else None,
                "debit": debit,
                "kredit": kredit,
                "saldo": saldo,
                "raw_text": raw_text,
            })
            prev_keterangan = keterangan

    return results


def _extract_text_fallback(
    file_path: str,
    bank_name: str | None = None,
    year: int | None = None,
) -> list[dict]:
    """Fallback: extract transactions from raw text when table extraction fails."""
    results = []
    row_num = 0

    # Pattern for a transaction line: starts with date, followed by description and amounts
    # DD/MM or DD/MM/YYYY ... amounts ...
    tx_pattern = re.compile(
        r"(\d{1,2}[/\-.]?\d{1,2}(?:[/\-.]?\d{2,4})?)"  # date
        r"\s+"
        r"(.+?)"  # description (non-greedy)
        r"\s+"
        r"([\d.,]+(?:\s*(?:DB|CR|Dr|Cr))?)"  # amount 1
        r"(?:\s+([\d.,]+(?:\s*(?:DB|CR|Dr|Cr))?))?"  # amount 2 (optional)
        r"(?:\s+([\d.,]+))?"  # amount 3 / saldo (optional)
    )

    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            for line in text.split("\n"):
                line = line.strip()
                if not line:
                    continue

                m = tx_pattern.match(line)
                if not m:
                    continue

                tanggal = _parse_date(m.group(1), year)
                if not tanggal:
                    continue

                keterangan = m.group(2).strip()
                amt1_raw = m.group(3)
                amt2_raw = m.group(4)
                amt3_raw = m.group(5)

                # Determine debit/credit from DB/CR markers
                debit = None
                kredit = None
                saldo = None

                if amt1_raw:
                    is_db = bool(re.search(r"(?i)db|dr", amt1_raw))
                    is_cr = bool(re.search(r"(?i)cr", amt1_raw))
                    amt1 = _parse_amount(re.sub(r"(?i)\s*(db|cr|dr)", "", amt1_raw))

                    if amt2_raw:
                        amt2 = _parse_amount(re.sub(r"(?i)\s*(db|cr|dr)", "", amt2_raw))
                        if amt3_raw:
                            saldo = _parse_amount(amt3_raw)
                        # Two amounts = debit and credit
                        debit = amt1
                        kredit = amt2
                    else:
                        # Single amount - check DB/CR marker
                        if is_db:
                            debit = amt1
                        elif is_cr:
                            kredit = amt1
                        else:
                            debit = amt1  # default to debit

                    if amt3_raw and not saldo:
                        saldo = _parse_amount(amt3_raw)

                row_num += 1
                results.append({
                    "row_number": row_num,
                    "tanggal": tanggal,
                    "keterangan": keterangan,
                    "debit": debit,
                    "kredit": kredit,
                    "saldo": saldo,
                    "raw_text": line,
                })

    return results


async def process_pdf(
    file_path: str,
    bank_name: str | None = None,
    year: int | None = None,
) -> list[dict]:
    """Extract and parse bank statement from PDF (runs in thread pool)."""
    def _do():
        # Try table-based extraction first
        tables = extract_tables_from_pdf(file_path, bank_name)
        results = parse_bank_statement(tables, bank_name, year)

        # If table extraction got very few results, try text fallback
        if len(results) < 3:
            logger.info(
                f"Table extraction got {len(results)} rows, trying text fallback..."
            )
            text_results = _extract_text_fallback(file_path, bank_name, year)
            if len(text_results) > len(results):
                logger.info(
                    f"Text fallback got {len(text_results)} rows (better), using it."
                )
                return text_results

        return results

    return await asyncio.to_thread(_do)


def get_total_pages(file_path: str) -> int:
    """Get total pages in a PDF."""
    with pdfplumber.open(file_path) as pdf:
        return len(pdf.pages)


def export_to_excel(
    results: list[dict],
    output_path: str,
    bank_name: str | None = None,
) -> str:
    """Export OCR results to a formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekening Koran"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    money_format = '#,##0.00'
    date_format = 'DD/MM/YYYY'

    # Title row
    title = f"Rekening Koran - {(bank_name or 'Bank').upper()}"
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells("A2:F2")
    ws["A2"].value = f"Diekspor pada: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=9, color="666666")
    ws.row_dimensions[2].height = 20

    # Headers
    headers = ["No", "Tanggal", "Keterangan", "Debit", "Kredit", "Saldo"]
    col_widths = [6, 14, 45, 18, 18, 18]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Data rows
    for i, row in enumerate(results, 1):
        r = i + 4
        ws.cell(row=r, column=1, value=i).border = thin_border
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")

        date_cell = ws.cell(row=r, column=2, value=row.get("tanggal"))
        date_cell.border = thin_border
        date_cell.number_format = date_format

        ws.cell(row=r, column=3, value=row.get("keterangan")).border = thin_border

        for col, key in [(4, "debit"), (5, "kredit"), (6, "saldo")]:
            val = row.get(key)
            cell = ws.cell(row=r, column=col, value=float(val) if val is not None else None)
            cell.border = thin_border
            cell.number_format = money_format
            cell.alignment = Alignment(horizontal="right")

    # Summary row
    if results:
        summary_row = len(results) + 5
        ws.cell(row=summary_row, column=3, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=3).border = thin_border

        for col in [4, 5]:
            cell = ws.cell(row=summary_row, column=col)
            start_cell = openpyxl.utils.get_column_letter(col) + "5"
            end_cell = openpyxl.utils.get_column_letter(col) + str(summary_row - 1)
            cell.value = f"=SUM({start_cell}:{end_cell})"
            cell.font = Font(bold=True)
            cell.number_format = money_format
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
