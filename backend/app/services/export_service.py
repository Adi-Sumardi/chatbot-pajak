"""Service for exporting chat AI responses to Excel and PDF files."""

import re
import logging
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from fpdf import FPDF

logger = logging.getLogger(__name__)


def _parse_markdown_tables(text: str) -> list[list[list[str]]]:
    """Extract markdown tables from text. Returns list of tables, each a list of rows."""
    tables = []
    lines = text.split("\n")
    current_table: list[list[str]] = []
    in_table = False

    for line in lines:
        stripped = line.strip()
        if "|" in stripped and stripped.startswith("|"):
            # Check if separator row (---|---|---)
            if re.match(r"^\|[\s\-:|]+\|$", stripped):
                continue
            cells = [c.strip() for c in stripped.split("|")[1:-1]]
            if cells:
                current_table.append(cells)
                in_table = True
        else:
            if in_table and current_table:
                tables.append(current_table)
                current_table = []
                in_table = False

    if current_table:
        tables.append(current_table)

    return tables


def _clean_markdown(text: str) -> list[dict]:
    """Parse markdown text into structured blocks for PDF rendering."""
    blocks = []
    lines = text.split("\n")
    i = 0

    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        if not stripped:
            blocks.append({"type": "space"})
            i += 1
            continue

        # Headings
        if stripped.startswith("### "):
            blocks.append({"type": "h3", "text": stripped[4:]})
        elif stripped.startswith("## "):
            blocks.append({"type": "h2", "text": stripped[3:]})
        elif stripped.startswith("# "):
            blocks.append({"type": "h1", "text": stripped[2:]})
        # Bullet list
        elif stripped.startswith("- ") or stripped.startswith("* "):
            blocks.append({"type": "bullet", "text": stripped[2:]})
        # Numbered list
        elif re.match(r"^\d+\.\s", stripped):
            text_content = re.sub(r"^\d+\.\s", "", stripped)
            blocks.append({"type": "numbered", "text": text_content})
        # Table row (collect whole table)
        elif "|" in stripped and stripped.startswith("|"):
            table_lines = []
            while i < len(lines) and "|" in lines[i].strip() and lines[i].strip().startswith("|"):
                row_stripped = lines[i].strip()
                if not re.match(r"^\|[\s\-:|]+\|$", row_stripped):
                    cells = [c.strip() for c in row_stripped.split("|")[1:-1]]
                    if cells:
                        table_lines.append(cells)
                i += 1
            if table_lines:
                blocks.append({"type": "table", "rows": table_lines})
            continue
        else:
            # Clean bold/italic markers for plain text
            clean = stripped.replace("**", "").replace("*", "").replace("`", "")
            blocks.append({"type": "text", "text": clean})

        i += 1

    return blocks


def export_chat_to_excel(content: str, output_path: str, title: str = "Laporan") -> str:
    """Export AI chat response to Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
    h2_font = Font(name="Calibri", bold=True, size=12, color="1E3A5F")
    bold_font = Font(name="Calibri", bold=True, size=10)
    normal_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtitle
    ws.merge_cells("A2:F2")
    ws["A2"].value = f"Dibuat oleh Chatbot Pajak - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=9, color="888888")

    # Parse content
    tables = _parse_markdown_tables(content)
    blocks = _clean_markdown(content)

    row = 4

    # If there are tables, write them formatted
    if tables:
        for table in tables:
            if not table:
                continue

            # Header row
            for col_idx, header in enumerate(table[0], 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            row += 1

            # Data rows
            for data_row in table[1:]:
                for col_idx, val in enumerate(data_row, 1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.font = normal_font
                    cell.border = thin_border
                    # Try to right-align numbers
                    if val and re.match(r"^[\d.,\-\s]+$", val.strip()):
                        cell.alignment = Alignment(horizontal="right")
                row += 1

            row += 1  # Gap between tables
    else:
        # No tables - write text content
        for block in blocks:
            if block["type"] == "space":
                row += 1
            elif block["type"] in ("h1", "h2", "h3"):
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                cell = ws.cell(row=row, column=1, value=block["text"])
                cell.font = h2_font if block["type"] in ("h1", "h2") else bold_font
                row += 1
            elif block["type"] in ("bullet", "numbered"):
                prefix = "• " if block["type"] == "bullet" else ""
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                ws.cell(row=row, column=1, value=f"  {prefix}{block['text']}").font = normal_font
                row += 1
            elif block["type"] == "table":
                for ri, trow in enumerate(block["rows"]):
                    for ci, val in enumerate(trow, 1):
                        cell = ws.cell(row=row, column=ci, value=val)
                        cell.border = thin_border
                        if ri == 0:
                            cell.font = header_font
                            cell.fill = header_fill
                        else:
                            cell.font = normal_font
                    row += 1
                row += 1
            elif block["type"] == "text":
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                ws.cell(row=row, column=1, value=block["text"]).font = normal_font
                row += 1

    # Auto-width columns
    for col in range(1, 7):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


class ChatPDF(FPDF):
    """Custom PDF with header and footer for chat exports."""

    def __init__(self, title: str = "Laporan"):
        super().__init__()
        self._doc_title = title

    def header(self):
        self.set_font("Helvetica", "B", 10)
        self.set_text_color(30, 58, 95)
        self.cell(0, 8, self._doc_title, new_x="LMARGIN", new_y="NEXT")
        self.set_draw_color(30, 58, 95)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(4)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 7)
        self.set_text_color(150, 150, 150)
        self.cell(0, 10, f"Chatbot Pajak - {datetime.now().strftime('%d/%m/%Y')} - Hal {self.page_no()}/{{nb}}", align="C")


def _strip_emoji(text: str) -> str:
    """Remove emoji and other non-latin1 characters that fpdf can't render."""
    return text.encode("latin-1", errors="ignore").decode("latin-1")


def export_chat_to_pdf(content: str, output_path: str, title: str = "Laporan") -> str:
    """Export AI chat response to PDF file."""
    content = _strip_emoji(content)
    title = _strip_emoji(title)
    pdf = ChatPDF(title=title)
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=20)

    blocks = _clean_markdown(content)

    for block in blocks:
        btype = block["type"]

        if btype == "space":
            pdf.ln(3)
        elif btype == "h1":
            pdf.set_font("Helvetica", "B", 16)
            pdf.set_text_color(30, 58, 95)
            pdf.cell(0, 10, block["text"], new_x="LMARGIN", new_y="NEXT")
            pdf.ln(2)
        elif btype == "h2":
            pdf.set_font("Helvetica", "B", 13)
            pdf.set_text_color(30, 58, 95)
            pdf.cell(0, 8, block["text"], new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)
        elif btype == "h3":
            pdf.set_font("Helvetica", "B", 11)
            pdf.set_text_color(30, 58, 95)
            pdf.cell(0, 7, block["text"], new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)
        elif btype == "bullet":
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(50, 50, 50)
            pdf.cell(8, 6, chr(8226))
            pdf.multi_cell(0, 6, block["text"], new_x="LMARGIN", new_y="NEXT")
        elif btype == "numbered":
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(50, 50, 50)
            pdf.multi_cell(0, 6, f"  {block['text']}", new_x="LMARGIN", new_y="NEXT")
        elif btype == "table":
            rows = block["rows"]
            if not rows:
                continue
            num_cols = max(len(r) for r in rows)
            col_width = (190) / max(num_cols, 1)

            for ri, trow in enumerate(rows):
                for ci in range(num_cols):
                    val = trow[ci] if ci < len(trow) else ""
                    if ri == 0:
                        pdf.set_font("Helvetica", "B", 9)
                        pdf.set_fill_color(30, 58, 95)
                        pdf.set_text_color(255, 255, 255)
                        pdf.cell(col_width, 7, val, border=1, fill=True, align="C")
                    else:
                        pdf.set_font("Helvetica", "", 9)
                        pdf.set_fill_color(255, 255, 255)
                        pdf.set_text_color(50, 50, 50)
                        pdf.cell(col_width, 6, val, border=1)
                pdf.ln()
            pdf.ln(3)
        elif btype == "text":
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(50, 50, 50)
            pdf.multi_cell(0, 6, block["text"], new_x="LMARGIN", new_y="NEXT")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    pdf.output(output_path)
    return output_path
